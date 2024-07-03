#********************************************************************
# System Imports
#********************************************************************
import sys
import imp
imp.reload(sys)
# sys.setdefaultencoding('utf8')
import os
import warnings

from random import uniform

#********************************************************************
# Own Imports
#********************************************************************
import shutil
import win32com.client
import openpyxl
from pyBase_path_cfg import *

#********************************************************************
# Constants
#********************************************************************
AP_TEST_CATALOGUE_NAME  = "Scenario_Catalogue_AP.xlsm"
# AP_TEST_CATALOGUE_NAME  = "Scenario_Catalogue_AP_Regression.xlsm"
AP_PARAM_NAME_FIRST_COL   = 1 # first column with parameter names in sheet Situations
AP_PARAM_NAME_ROW         = 7 # row with parameter names in sheet Situations
AP_FIRST_TEST_ROW         = 8 # first Scenario Test row in sheet Situations
AP_VEH_TYPE_COL_NUM       = 1 # column for vehicle name in sheet Vehicles
AP_VEH_WIDTH_COL_NUM      = 3 # column for vehicle width parameter in sheet Vehicles
AP_VEH_LENGTH_COL_NUM     = 4 # column for vehicle length parameter in sheet Vehicles
AP_VEH_OVERHANG_COL_NUM   = 6 # column for vehicle overhang parameter in sheet Vehicles
AP_VEH_WHEELBASE_COL_NUM  = 7 # column for vehicle wheelbase parameter in sheet Vehicles
AP_LIM_TYPE_COL_NUM       = 1 # column for limiter type in sheet Types
AP_LIM_HEIGHT_COL_NUM     = 3 # column for limiter height parameter in sheet Types
AP_BUMP_TYPE_COL_NUM      = 1 # column for bump type in sheet Types
AP_BUMP_HEIGHT_COL_NUM    = 3 # column for bump height parameter in sheet Types
AP_BUMP_RAMPUP_COL_NUM    = 4 # column for bump ramp up length parameter in sheet Types
AP_BUMP_PLATEAU_COL_NUM   = 5 # column for bump plateau length parameter in sheet Types
AP_BUMP_RAMPDOWN_COL_NUM  = 6 # column for bump ramp down length parameter in sheet Types
AP_WHLSTP_TYPE_COL_NUM    = 1 # column for wheelstopper type in sheet Types
AP_WHLSTP_HEIGHT_COL_NUM  = 3 # column for wheelstopper height in sheet Types
AP_WHLSTP_LENGTH_COL_NUM  = 4 # column for wheelstopper length in sheet Types
AP_WHLSTP_WIDTH_COL_NUM   = 5 # column for wheelstopper width in sheet Types

# TODO: verify if type attributes should be switched from dropdown to list
# type: type_Ego, type_P1, type_P2, slanted, lim_Rd_P, lim_P1_P, lim_P2_P, surface_P, surface_R, shape_Obs1, area_Obs1, shape_Obs2, area_Obs2, shape_Obs3, area_Obs3

#********************************************************************
# Configuration tables
#********************************************************************
limiterConfig = [
    "curb_t",
    "curb_tb",
    "curb_ntd",
    "curb_nt",
    "guard",
    "wall"
    ]
    
bumpConfig = [
    "speed_bump",
    "inv_speed_bump",
    "speed_hump"
    ]
    
wheelstopperConfig = [
    "single_beam"
    ]
	
class AP_ExcelReader:
    # Class constructor
    def __init__(self):
        # Load the Workbook
        print("Loading Excel-file...")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            book = openpyxl.load_workbook(os.path.join(IMS_PATH, AP_TEST_CATALOGUE_NAME), data_only = True)
        print("done")
        
        self.sheetForSituations = book['Situations']
        self.sheetForVehicles = book['Vehicles']
        self.sheetForTypes = book['Types']
        
        self.firstTestRow = AP_FIRST_TEST_ROW
        self.lastTestRow = self.sheetForSituations.max_row # last ST row in sheet Situations
        
        # Read parameter names from test catalogue
        self.paramConfig = []
        for col in range(AP_PARAM_NAME_FIRST_COL, self.sheetForSituations.max_column + 1):
            colName = self.sheetForSituations.cell(row = AP_PARAM_NAME_ROW, column = col).value
            self.paramConfig.append(colName)
        
        # Height of curbstone (m)
        self.limHeight = {}
        for limiterName in limiterConfig:
            self.limHeight[limiterName] =  self.getLimiterData(limiterName)
            
        # Parameter info for bumps
        self.bumpHeight = {}
        self.bumpRampUp = {}
        self.bumpPlateau = {}
        self.bumpRampDown = {}
        for bumpName in bumpConfig:
            (self.bumpHeight[bumpName], self.bumpRampUp[bumpName], self.bumpPlateau[bumpName], self.bumpRampDown[bumpName]) = self.getBumpData(bumpName)
            
        # Parameter info for wheelstoppers
        self.wheelstopperHeight = {}
        self.wheelstopperLength = {}
        self.wheelstopperWidth = {}
        for wheelstopperName in wheelstopperConfig:
            (self.wheelstopperHeight[wheelstopperName], self.wheelstopperLength[wheelstopperName], self.wheelstopperWidth[wheelstopperName]) = self.getWheelstopperData(wheelstopperName)
        
        self.rowNumValues = 0 # number of test cases in a row
        
    # Use with AP TestCatalog (TestRun excel file)
    def extractScenarios(self, currRow):
        paramMultiValue = {} # dictionary of (pos['parameter name'] <-> list of values for 'parameter name')
        scenario = {} # dictionary of ('parameter name' <-> parameter value for CarMaker scenario)
        scenarios = [] # list of dictionaries (one list entry contains concrete parameter values for one scenario)
        # pos is required for cartesian product of paramMultiValue
        self.rowNumValues = 0

        for paramName in self.paramConfig:
            paramMultiValue[self.paramConfig.index(paramName)] = self.readCell(currRow, paramName)
            
        if self.rowNumValues > 1:
            for paramName in self.paramConfig:
                if len(paramMultiValue[self.paramConfig.index(paramName)]) == 1:
                    paramMultiValue[self.paramConfig.index(paramName)] = paramMultiValue[self.paramConfig.index(paramName)] * self.rowNumValues # copy the same value for 'rowNumValues' times

        # izip returns an interator for zip
        # * unpacks argumet lists
        # zip* returns a list of tuples - e.g. [[A, B, C], [1, 2, 3]] => [(A,1), (B,2), (C,3)]
        # zip without * would return [([A, B, C],), ([1, 2, 3],)]
        for currSetup in zip(*iter(paramMultiValue.values())): 
            scenario = {}
            for paramName in self.paramConfig:
                scenario[paramName] = currSetup[self.paramConfig.index(paramName)]
                
            vehData = self.getVehicleData(scenario["type_Ego"]) # get vehicle data for ego vehicle
            scenario["w_Ego"] = vehData["width"] 
            scenario["l_Ego"] = vehData["length"]
            scenario["o_Ego"] = vehData["overhang"]
            scenario["wb_Ego"] = vehData["wheelbase"]
            scenario["h_Ego"] = vehData["height"]
            
            if scenario["exists_P1"] == "true":
                vehData = self.getVehicleData(scenario["type_P1"]) # get vehicle data for parking vehicle 1
                scenario["w_P1"] = vehData["width"] 
                scenario["l_P1"] = vehData["length"]
                scenario["o_P1"] = vehData["overhang"]
                scenario["wb_P1"] = vehData["wheelbase"]
                scenario["h_P1"] = vehData["height"]
                if scenario["Category"].find("Perpendicular") != -1 or scenario["Category"].find("Angled") != -1:
                    scenario["psi_P1"] = self.getRotation(scenario["d_P1_C1"], scenario["d_P1_C2"], scenario["l_P1"], scenario["Category"])
                elif scenario["Category"].find("Parallel") != -1:
                    scenario["psi_P1"] = self.getRotation(scenario["d_P1_C1"], scenario["d_P1_C2"], scenario["w_P1"], scenario["Category"])
                else:
                    scenario["psi_P1"] = 0
                scenario["psi_P1"] *= -1 # the rotation is flipped for P1 because d_P1_Cx is flipped compared to P2
                
            if scenario["exists_P2"] == "true":
                vehData = self.getVehicleData(scenario["type_P2"]) # get vehicle data for parking vehicle 2
                scenario["w_P2"] = vehData["width"] 
                scenario["l_P2"] = vehData["length"]
                scenario["o_P2"] = vehData["overhang"]
                scenario["wb_P2"] = vehData["wheelbase"]
                scenario["h_P2"] = vehData["height"]
                if scenario["Category"].find("Perpendicular") != -1 or scenario["Category"].find("Angled") != -1:
                    scenario["psi_P2"] = self.getRotation(scenario["d_P2_C1"], scenario["d_P2_C2"], scenario["l_P2"], scenario["Category"])
                elif scenario["Category"].find("Parallel") != -1:
                    scenario["psi_P2"] = self.getRotation(scenario["d_P2_C1"], scenario["d_P2_C2"], scenario["w_P2"], scenario["Category"])
                else:
                    scenario["psi_P2"] = 0
            
            scenarios.append(scenario)
        
        return scenarios
       
    def getRotation(self, d_Px_C1, d_Px_C2, side_Size, category):
        # Get differences Px_C1 corner minus Px_C2 corner (left for P1, right for P2)
        diffs = [0, 0]
        diffs[0] = d_Px_C1 - d_Px_C2
        if category.find("Left") != -1:
            diffs[0] *= -1 # this seems to work (don't know yet why)
        diffs[1] = sqrt(side_Size**2 - diffs[0]**2) # Pythagorean theorem within right triangle with diffs[0], diffs[1] legs and side_Size hypotenuse
        
        # Get rotation in degrees
        if  diffs[1] != 0:
            rotation = atan(diffs[0]/diffs[1]) * 180 / pi
        else:
            rotation = 0

        # # Adjust for 2nd & 3rd quadrants, i.e. diff y is -ve.
        # if (diffs[1] < 0) :
            # rotation += 180

        # # Adjust for 4th quadrant
        # # i.e. diff x is -ve, diff y is +ve
        # elif(diffs[0] < 0) :
            # rotation += 360

        # return array of rotation
        return rotation
        
    def readCell(self, currRow, paramName):
        retValueList = []
        tempCellValue = str(self.sheetForSituations.cell(row = currRow, column = self.paramConfig.index(paramName) + 1).value)
        # tempCellValue = self.sheetForSituations.cell(row = currRow, column = self.col[paramName]).value
        # tempCellValue = tempCellValue.encode("utf-8")
        
        if tempCellValue.find(';') != - 1: # multiple values cell
            tempCellValue.replace(' ', '') # remove blanks from cell
            retValueList = tempCellValue.split(";") # extract enumeration
            if self.rowNumValues <= 1:
                self.rowNumValues = len(retValueList) # this is the first cell with multiple values
            elif self.rowNumValues != len(retValueList): # number of values in current cell different than expected
                print(("Number of values for {0} on row {1} is different than {2}".format(paramName, currRow, self.rowNumValues)))
                sys.exit(0)
        else:
            retValueList = [tempCellValue]
            if self.rowNumValues == 0: # no multiple values cells on currRow until now
                self.rowNumValues = 1
        try:
            retValueList = list(map(float, retValueList))
        except Exception as e:
            pass
        return retValueList
    
    def getColNumByName(self, string, sheet, rowToSearch):
        colPos = -1
        colCount = sheet.max_column
        for i in range (1, colCount + 1):
            colName = sheet.cell(row = rowToSearch, column = i).value
            if colName != [] and str(string) == str(colName):
                colPos = i
        return colPos
        
    def getRowNumByName(self, string, sheet, colToSearch):
        rowPos = -1
        rowCount = sheet.max_row
        for i in range (1, rowCount + 1):
            rowName = sheet.cell(row = i, column = colToSearch).value
            if rowName != [] and str(string) == str(rowName):
                rowPos = i
        return rowPos
        
    def getVehicleData(self, carType):
        # vehicles parameters
        rowVehicle = self.getRowNumByName(carType, self.sheetForVehicles, AP_VEH_TYPE_COL_NUM)
        vehData = {}
        
        if rowVehicle == -1:
            print(("Car type {} does not exist in sheet Vehicles".format(carType)))
            sys.exit(0)
        else:
            vehData["width"] = self.sheetForVehicles.cell(row = rowVehicle, column = AP_VEH_WIDTH_COL_NUM).value
            vehData["length"] = self.sheetForVehicles.cell(row = rowVehicle, column = AP_VEH_LENGTH_COL_NUM).value
            vehData["overhang"] = self.sheetForVehicles.cell(row = rowVehicle, column = AP_VEH_OVERHANG_COL_NUM).value
            vehData["wheelbase"] = self.sheetForVehicles.cell(row = rowVehicle, column = AP_VEH_WHEELBASE_COL_NUM).value
            vehData["height"] = 1.5 # default value used for CarMaker collision box
            
        return vehData
        
    def getLimiterData(self, limiterType):
        # limiter parameters
        rowLimiter = self.getRowNumByName(limiterType, self.sheetForTypes, AP_LIM_TYPE_COL_NUM)
        limiterData = 0
        
        if rowLimiter == -1:
            print(("Limiter type {} does not exist in sheet Types".format(limiterType)))
            sys.exit(0)
        else:
            limiterData = self.sheetForTypes.cell(row = rowLimiter, column = AP_LIM_HEIGHT_COL_NUM).value
            
        return limiterData
        
    def getBumpData(self, bumpType):
        # bump parameters
        rowBump = self.getRowNumByName(bumpType, self.sheetForTypes, AP_BUMP_TYPE_COL_NUM)
        
        if rowBump == -1:
            print(("Bump type {} does not exist in sheet Types".format(bumpType)))
            sys.exit(0)
        else:
            bumpHeight = self.sheetForTypes.cell(row = rowBump, column = AP_BUMP_HEIGHT_COL_NUM).value
            bumpRampUp = self.sheetForTypes.cell(row = rowBump, column = AP_BUMP_RAMPUP_COL_NUM).value
            bumpPlateau = self.sheetForTypes.cell(row = rowBump, column = AP_BUMP_PLATEAU_COL_NUM).value
            bumpRampDown = self.sheetForTypes.cell(row = rowBump, column = AP_BUMP_RAMPDOWN_COL_NUM).value
            
        return (bumpHeight, bumpRampUp, bumpPlateau, bumpRampDown)
        
    def getWheelstopperData(self, wheelstopperType):
        # wheelstopper parameters
        rowWheelstopper = self.getRowNumByName(wheelstopperType, self.sheetForTypes, AP_WHLSTP_TYPE_COL_NUM)
        
        if rowWheelstopper == -1:
            print(("Wheelstopper type {} does not exist in sheet Types".format(wheelstopperType)))
            sys.exit(0)
        else:
            wheelstopperHeight = self.sheetForTypes.cell(row = rowWheelstopper, column = AP_WHLSTP_HEIGHT_COL_NUM).value
            wheelstopperLength = self.sheetForTypes.cell(row = rowWheelstopper, column = AP_WHLSTP_LENGTH_COL_NUM).value
            wheelstopperWidth = self.sheetForTypes.cell(row = rowWheelstopper, column = AP_WHLSTP_WIDTH_COL_NUM).value
            
        return (wheelstopperHeight, wheelstopperLength, wheelstopperWidth)

    def existsScenario(self, currRow):
        exists = False
        if self.sheetForSituations.cell(row = currRow, column = self.paramConfig.index("Situation ID") + 1).value != None: # the row contains a test situation
            exists = True
            
        return exists
        
    def toBeGenerated(self, currRow):
        toBeGen = False
        if self.sheetForSituations.cell(row = currRow, column = self.paramConfig.index("Generate Testcase") + 1).value == "yes": # the test situation must be generated
            toBeGen = True

        return toBeGen
        