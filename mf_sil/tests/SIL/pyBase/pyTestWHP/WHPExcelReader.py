#********************************************************************
# System Imports
#********************************************************************
import sys
import imp

imp.reload(sys)
# sys.setdefaultencoding('utf8')
import os
import warnings
import re
from random import uniform

#********************************************************************
# Own Imports
#********************************************************************
import shutil
import win32com.client
import openpyxl
from pyBase_path_cfg import *
from collections import namedtuple

#********************************************************************
# Constants
#********************************************************************
WHP_PARAM_NAME_ROW       = 1 # first row with parameter names
WHP_FIRST_TEST_ROW       = 2 # first Scenario Test row in sheet Situations
WHP_FIRST_BTN_ROW        = 2 # first Button row in sheet HMI
WHP_FIRST_VHCL_ROW       = 2 # first Vehicle row in sheet Vehicles
WHP_FIRST_STRFURN_ROW    = 2 # first Furniture row in sheet StreetFurniture
WHP_FIRST_PARAM_ROW      = 2 # first Parameter row in sheet Parameters

#********************************************************************
# Configuration tables
#********************************************************************

paramConfig = [
    "Situation ID",
    "Category",
    "Description",
    "Generate Testcase",
    "Start Long Offset",
    "Start Lat Offset",
    "Start Angle",
    "Changed parameters",
    "Maneuvers",
    "Obstacles",
    "Evaluation"
    ]

vehicleParamConfigType = namedtuple("vehicleParamConfigType", "length width height movieGeometry")
strFurnParamConfigType = namedtuple("strFurnParamConfigType", "length width height movieGeometry")
whpParamConfigType = namedtuple("whpParamConfigType", "value factor")

class WHPExcelReader:
    # Class constructor
    def __init__(self):
        # Parameters sheet needs writing, so open catalogue with data_only = False before saving it, in order to preserve formulas
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            bookToSave = openpyxl.load_workbook(TESTCATALOGUE_PATH, data_only = False, keep_vba = True)
            
        self.sheetForParameters = bookToSave['Parameters']
        self.Parameters = {}
        colKey = self.getColNumByName("Parameter name", self.sheetForParameters, WHP_PARAM_NAME_ROW)
        colFactor = self.getColNumByName("Desired factor", self.sheetForParameters, WHP_PARAM_NAME_ROW)
        colValue = self.getColNumByName("Parameter value - DO NOT FILL", self.sheetForParameters, WHP_PARAM_NAME_ROW)
        for currRow in range(WHP_FIRST_PARAM_ROW, self.sheetForParameters.max_row + 1):
            key = str(self.sheetForParameters.cell(row = currRow, column = colKey).value)
            if key == "None":
                continue
            if self.sheetForParameters.cell(row = currRow, column = colFactor).value:
                try:
                    factor = float(self.sheetForParameters.cell(row = currRow, column = colFactor).value)
                except ValueError:
                    print("Unexpected factor value in line {}.".format(currRow))
                    sys.exit(0)
            else:
                factor = 1
            value = self.getParamValueFromSWBuild(key)
            self.sheetForParameters.cell(row = currRow, column = colValue).value = value
            self.Parameters[key] = whpParamConfigType(value, factor)          
        bookToSave.save(TESTCATALOGUE_PATH)
        bookToSave =  None

        # The rest of the sheets need to be read with formulas evaluated, so open workbook with data_only = True
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            book = openpyxl.load_workbook(TESTCATALOGUE_PATH, data_only = True, keep_vba = True)

        self.sheetForSituations = book['Situations']
        self.sheetForHMI = book['HMI']
        self.sheetForVehicles = book['Vehicles']
        self.sheetForStreetFurniture = book['StreetFurniture']

        self.firstTestRow = WHP_FIRST_TEST_ROW
        self.lastTestRow = self.sheetForSituations.max_row # last ST row in sheet Situations

        self.col = {}
        # fill out col[paramName] = column number in sheet "Situations"
        for paramName in paramConfig:
            colParam = self.getColNumByName(paramName, self.sheetForSituations, WHP_PARAM_NAME_ROW)
            if colParam == -1: # search for parameter name in row
                print(("There is no column for parameter: {} in sheet Situations".format(paramName)))
                sys.exit(0) # stop script execution
            else:
                self.col[paramName] = colParam
        
        self.HMIButtons = {}
        colKey = self.getColNumByName("Button", self.sheetForHMI, WHP_PARAM_NAME_ROW)
        colValue = self.getColNumByName("Value", self.sheetForHMI, WHP_PARAM_NAME_ROW)
        for currRow in range(WHP_FIRST_BTN_ROW, self.sheetForHMI.max_row + 1):
            key = str(self.sheetForHMI.cell(row = currRow, column = colKey).value)
            value = str(self.sheetForHMI.cell(row = currRow, column = colValue).value)
            value = value.replace(" ","") # remove blanks from cell
            self.HMIButtons[key] = value
        
        self.Vehicles = {}
        colKey = self.getColNumByName("Model", self.sheetForVehicles, WHP_PARAM_NAME_ROW)
        colLength = self.getColNumByName("Length", self.sheetForVehicles, WHP_PARAM_NAME_ROW)
        colWidth = self.getColNumByName("Width", self.sheetForVehicles, WHP_PARAM_NAME_ROW)
        colHeight = self.getColNumByName("Height", self.sheetForVehicles, WHP_PARAM_NAME_ROW)
        colMovieGeometry = self.getColNumByName("Movie geometry", self.sheetForVehicles, WHP_PARAM_NAME_ROW)
        for currRow in range(WHP_FIRST_VHCL_ROW, self.sheetForVehicles.max_row + 1):
            key = str(self.sheetForVehicles.cell(row = currRow, column = colKey).value)
            length = float(self.sheetForVehicles.cell(row = currRow, column = colLength).value)
            width = float(self.sheetForVehicles.cell(row = currRow, column = colWidth).value)
            height = float(self.sheetForVehicles.cell(row = currRow, column = colHeight).value)
            movieGeometry = str(self.sheetForVehicles.cell(row = currRow, column = colMovieGeometry).value)
            self.Vehicles[key] = vehicleParamConfigType(length, width, height, movieGeometry)

        self.StrFurniture = {}
        colKey = self.getColNumByName("Type", self.sheetForStreetFurniture, WHP_PARAM_NAME_ROW)
        colLength = self.getColNumByName("Length", self.sheetForStreetFurniture, WHP_PARAM_NAME_ROW)
        colWidth = self.getColNumByName("Width", self.sheetForStreetFurniture, WHP_PARAM_NAME_ROW)
        colHeight = self.getColNumByName("Height", self.sheetForStreetFurniture, WHP_PARAM_NAME_ROW)
        colMovieGeometry = self.getColNumByName("Movie geometry", self.sheetForStreetFurniture, WHP_PARAM_NAME_ROW)
        for currRow in range(WHP_FIRST_STRFURN_ROW, self.sheetForStreetFurniture.max_row + 1):
            key = str(self.sheetForStreetFurniture.cell(row = currRow, column = colKey).value)
            length = float(self.sheetForStreetFurniture.cell(row = currRow, column = colLength).value)
            width = float(self.sheetForStreetFurniture.cell(row = currRow, column = colWidth).value)
            height = float(self.sheetForStreetFurniture.cell(row = currRow, column = colHeight).value)
            movieGeometry = str(self.sheetForStreetFurniture.cell(row = currRow, column = colMovieGeometry).value)
            self.StrFurniture[key] = strFurnParamConfigType(length, width, height, movieGeometry)
        
        
    # Use with WHP TestCatalog (TestRun excel file)
    def extractScenario(self, currRow):
        scenario = {} # dictionary of ('parameter name' <-> parameter value for CarMaker scenario)

        for paramName in paramConfig:
            scenario[paramName] = str(self.sheetForSituations.cell(row = currRow, column = self.col[paramName]).value)

        return scenario

    def getColNumByName(self, string, sheet, rowToSearch):
        colPos = -1
        colCount = sheet.max_column
        for i in range (1, colCount + 1):
            colName = sheet.cell(row = rowToSearch, column = i).value
            if colName != [] and str(string) == str(colName):
                colPos = i
        return colPos

    def getRowNumByName(self, string, sheet, colToSearch):
        rowPos = -1
        rowCount = sheet.max_row
        for i in range (1, rowCount + 1):
            rowName = sheet.cell(row = i, column = colToSearch).value
            if rowName != [] and str(string) == str(rowName):
                rowPos = i
        return rowPos

    def existsScenario(self, currRow):
        exists = False
        colSituation = self.getColNumByName("Situation ID", self.sheetForSituations, WHP_PARAM_NAME_ROW)
        if colSituation != -1:
            if self.sheetForSituations.cell(row = currRow, column = colSituation).value != None: # the row contains a test situation
                exists = True
        else:
            print("Situation ID not found.")
            sys.exit(0)
        return exists

    def toBeGenerated(self, currRow):
        toBeGen = False
        colGen = self.getColNumByName("Generate Testcase", self.sheetForSituations, WHP_PARAM_NAME_ROW)
        if colGen != -1:
            if self.sheetForSituations.cell(row = currRow, column = colGen).value == "yes": # the row contains a test situation
                toBeGen = True
        else:
            print("Generate Testcase column not found.")
            sys.exit(0)
        return toBeGen

    def getParamValueFromSWBuild(self, paramName):
        reParamLine = r"\"" + paramName + "\":\s*(?P<paramVal>[\d.-]*)" # json format for one parameter is "PARAM_NAME": value (create named group for value)
        for path in PARAM_FILES_LIST:
            with open(path, 'r') as jsonFile:
                jsonText = jsonFile.read()
                found = re.search(reParamLine, jsonText)
                if found:
                    return float(found.group('paramVal'))
        print("Parameter {} from Parameters sheet not found in SW build. Please check if name is correct or if list of json files from pyBase_path_cfg.py must be extended.".format(paramName))
        sys.exit(0)
