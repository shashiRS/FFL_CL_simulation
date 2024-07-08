import sys
import re
import os

import openpyxl

from decimal import *

MAX_NUMBER_OF_DIGITS = 7

def get_column_values(wb_obj,column_names):
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    result = {}
    column_numbers = {}

    for column in column_names:
        result[column] = []
        for i in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=1, column=i)
            if cell_obj.value == column:
                column_numbers[column] = i
  
    for i in range(2, max_row + 1):
        no_NA = True
        for column in column_numbers.keys():
            cell_obj = sheet_obj.cell(row=i, column=column_numbers[column])
            if cell_obj.value in ['NA', 'TBD', ''] or '[Obsolete]' in str(cell_obj.value) or (column == '_Value' and isinstance(cell_obj.value, str) and re.match(r"[a-zA-Z ]+", cell_obj.value)):
                no_NA = False
                break
        if no_NA:
            for column in column_numbers.keys():
                cell_obj = sheet_obj.cell(row=i, column=column_numbers[column])
                if column == '_Value':
                    digitsNumber = 0
                    value = cell_obj.value
                    #check for scientific notation and calculate the number of digits
                    if re.search("[0-9]e-",str(value)):
                        digitsNumber = (len(str(value).split('e-')[0].replace('.','')) + int(str(value).split('e-')[1])) if digitsNumber < MAX_NUMBER_OF_DIGITS else MAX_NUMBER_OF_DIGITS
                        value = round(Decimal(value),digitsNumber)
                    #check if the number is a float and calculate the number of digits
                    elif isinstance(value, float):
                        digitsNumber = int(str(Decimal(str(value)).as_tuple().exponent).lstrip('-')) if digitsNumber < MAX_NUMBER_OF_DIGITS else MAX_NUMBER_OF_DIGITS
                        value = round(Decimal(value),digitsNumber)

                    result[column].append(value)
                else:
                    result[column].append(cell_obj.value)
    wb_obj.close()
    return result

def create_params_dict(params):
    result = {}
    listOfParamName = []
    listOfParamValue = []
    for name in params['_ParameterName']:
        listOfParamName.append(name)

    for value in params['_Value']:
        listOfParamValue.append(value)
    
    if len(listOfParamName) == len(listOfParamValue):
        for idx in range(0,len(listOfParamName)):
            result[listOfParamName[idx]] = listOfParamValue[idx]

    return result

def write_param_to_file(out_path,params):
    try:
        out_file = open(out_path, 'w')
        lines = ''
        for key in params.keys():
            lines = lines + key + ' ' + str(params[key]) + '\n'
        out_file.write(lines)
        out_file.close()
    except Exception as e:
        print(e)
        return False
    return True

def readParamsFromFile(in_path):
    result = {}
    with open(in_path) as f:
        content = f.readlines()
    content = [x.strip("\n") for x in content]
    for x in content:
        res = x.split(" ")
        result[res[0]] = res[1]
    return result

if __name__ == "__main__":
    EVALPATH = os.path.dirname(__file__)
    in_excel_path         = os.path.abspath(os.path.join(EVALPATH, 'DOORS_Export_AUP_Parameters_list.xlsx'))
    additional_param_path = os.path.abspath(os.path.join(EVALPATH, 'parameters_notFoundInDoors.txt'))
    out_txt_path          = os.path.abspath(os.path.join(EVALPATH, '..', 'ImportParametersList_AUP.txt'))

    try:
        wb_obj = openpyxl.load_workbook(in_excel_path)
        params = get_column_values(wb_obj,['_ParameterName','_Value'])
        params = create_params_dict(params)
        additional_params = readParamsFromFile(additional_param_path)
        params.update(additional_params)
        write_param_to_file(out_txt_path, params)
        os.remove(in_excel_path)
    except Exception as e:
        print(e)
    